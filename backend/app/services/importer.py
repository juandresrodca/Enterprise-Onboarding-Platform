"""Bulk import: CSV, Excel (.xlsx) and JSON to NewUserSpec lists.

Headers are normalized and matched against a generous alias map so exports
from HR systems (Workday, SAP HR, BambooHR) import without manual renaming.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import date, datetime
from typing import Any

from pydantic import ValidationError

from app.core.exceptions import ValidationFailed
from app.models.user import NewUserSpec, ValidationIssue

# canonical field -> accepted header aliases (normalized: lowercase, [a-z0-9_])
_ALIASES: dict[str, list[str]] = {
    "first_name": ["first_name", "firstname", "given_name", "givenname", "first"],
    "last_name": ["last_name", "lastname", "surname", "family_name", "last"],
    "display_name": ["display_name", "displayname", "full_name"],
    "sam_account_name": ["sam_account_name", "samaccountname", "username", "sam", "login"],
    "user_principal_name": ["user_principal_name", "userprincipalname", "upn"],
    "email": ["email", "mail", "email_address", "primary_smtp"],
    "ou": ["ou", "organizational_unit", "organizationalunit", "target_ou", "container"],
    "department": ["department", "dept"],
    "company": ["company", "organization", "organisation"],
    "office": ["office", "physical_office"],
    "office_location": ["office_location", "location", "site"],
    "job_title": ["job_title", "title", "jobtitle", "position"],
    "employee_id": ["employee_id", "employeeid", "emp_id", "staff_id"],
    "employee_type": ["employee_type", "employeetype", "worker_type"],
    "cost_center": ["cost_center", "costcenter", "cc"],
    "description": ["description", "notes"],
    "manager": ["manager", "manager_sam", "manager_upn", "reports_to"],
    "phone": ["phone", "office_phone", "telephone", "telephonenumber"],
    "mobile": ["mobile", "mobile_phone", "cell", "cellphone"],
    "country": ["country", "country_code", "co"],
    "city": ["city", "town", "l"],
    "state": ["state", "province", "st"],
    "address": ["address", "street", "street_address", "streetaddress"],
    "postal_code": ["postal_code", "postalcode", "zip", "zip_code"],
    "account_expiration": ["account_expiration", "expiration", "expiry_date", "end_date"],
    "groups": ["groups", "group_memberships", "memberof"],
    "licenses": ["licenses", "license", "license_type", "skus"],
    "create_mailbox": ["create_mailbox", "mailbox"],
    "shared_mailboxes": ["shared_mailboxes", "sharedmailboxes"],
    "home_folder_enabled": ["home_folder_enabled", "home_folder", "homefolder"],
    "home_drive": ["home_drive", "home_drive_letter", "homedrive"],
    "logon_script": ["logon_script", "logonscript", "script_path"],
    "roaming_profile_path": ["roaming_profile_path", "profile_path", "profilepath"],
    "password": ["password", "initial_password", "temp_password"],
}

_HEADER_LOOKUP = {alias: field for field, aliases in _ALIASES.items() for alias in aliases}

TEMPLATE_HEADERS = [
    "first_name", "last_name", "sam_account_name", "ou", "department", "company",
    "job_title", "employee_id", "employee_type", "cost_center", "manager", "office",
    "office_location", "phone", "mobile", "country", "city", "state", "address",
    "postal_code", "account_expiration", "groups", "licenses", "create_mailbox",
    "shared_mailboxes", "home_folder_enabled", "home_drive", "logon_script",
    "description",
]

_LIST_SPLIT = re.compile(r"[;|]")
_TRUTHY = {"1", "true", "yes", "y", "x"}


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(header).strip().lower()).strip("_")


def _to_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [p.strip() for p in _LIST_SPLIT.split(str(value)) if p.strip()]


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in _TRUTHY


def _to_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text  # let pydantic raise a precise error


def rows_to_specs(
    rows: list[dict[str, Any]],
) -> tuple[list[NewUserSpec], list[ValidationIssue]]:
    users: list[NewUserSpec] = []
    issues: list[ValidationIssue] = []
    for i, raw in enumerate(rows):
        mapped: dict[str, Any] = {}
        for key, value in raw.items():
            field = _HEADER_LOOKUP.get(_normalize_header(key))
            if field is None or value in (None, ""):
                continue
            mapped[field] = value

        payload: dict[str, Any] = {
            k: v for k, v in mapped.items()
            if k in NewUserSpec.model_fields and k not in (
                "groups", "licenses", "shared_mailboxes", "account_expiration",
                "create_mailbox",
            )
        }
        payload["groups"] = _to_list(mapped.get("groups"))
        payload["licenses"] = _to_list(mapped.get("licenses"))
        payload["shared_mailboxes"] = _to_list(mapped.get("shared_mailboxes"))
        payload["create_mailbox"] = _to_bool(mapped.get("create_mailbox"), default=True)
        payload["account_expiration"] = _to_date(mapped.get("account_expiration"))
        payload["home_folder"] = {
            "enabled": _to_bool(mapped.get("home_folder_enabled")),
            "drive_letter": str(mapped.get("home_drive") or "H").strip(":").upper() or "H",
        }
        payload["profile"] = {
            "roaming_profile_path": mapped.get("roaming_profile_path"),
            "logon_script": mapped.get("logon_script"),
        }
        if mapped.get("password"):
            payload["password"] = {"generate": False, "value": str(mapped["password"])}

        try:
            users.append(NewUserSpec(**payload))
        except ValidationError as exc:
            for err in exc.errors():
                issues.append(ValidationIssue(
                    index=i,
                    field=".".join(str(p) for p in err["loc"]),
                    code="parse_error",
                    severity="error",
                    message=f"Row {i + 2}: {err['msg']}",
                ))
    return users, issues


def parse_upload(filename: str, content: bytes) -> list[dict[str, Any]]:
    """Dispatch on file extension; returns raw row dicts."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(content)
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    if name.endswith(".json"):
        return _parse_json(content)
    raise ValidationFailed(
        f"Unsupported file type: {filename}. Use .csv, .xlsx or .json"
    )


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Sniff the delimiter from the header line only; data cells may legally
    # contain semicolons (multi-value fields) or commas (quoted DNs).
    header_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = ";" if header_line.count(";") > header_line.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValidationFailed("CSV file has no header row")
    return [row for row in reader if any((v or "").strip() for v in row.values())]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h or "") for h in next(rows)]
    except StopIteration:
        raise ValidationFailed("Excel sheet is empty") from None
    out = []
    for values in rows:
        if values is None or all(v in (None, "") for v in values):
            continue
        out.append(dict(zip(headers, values)))
    wb.close()
    return out


def _parse_json(content: bytes) -> list[dict[str, Any]]:
    try:
        data = json.loads(content.decode("utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise ValidationFailed(f"Invalid JSON: {exc}") from None
    if isinstance(data, dict) and isinstance(data.get("users"), list):
        data = data["users"]
    if not isinstance(data, list):
        raise ValidationFailed("JSON must be an array of user objects (or {\"users\": [...]})")
    return data


def build_template_csv() -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow([
        "Jane", "Doe", "", "OU=Finance,OU=Company,DC=northwind,DC=local", "Finance",
        "Northwind Dynamics", "Financial Analyst", "EMP-2001", "Employee", "CC-FIN-01",
        "john.smith", "Seattle HQ", "Seattle - Floor 2", "+1 206 555 0100",
        "+1 425 555 0100", "US", "Seattle", "WA", "700 Rainier Ave", "98101",
        "", "SG-Finance-Users;DL-Finance", "SPE_E3", "true", "", "true", "H",
        "logon.bat", "Financial Analyst, Finance",
    ])
    writer.writerow([
        "Max", "Mustermann", "", "OU=IT,OU=Company,DC=northwind,DC=local", "IT",
        "Northwind Dynamics", "Systems Engineer", "EMP-2002", "Employee", "CC-IT-01",
        "elena.moreau", "Madrid Office", "Madrid - Floor 1", "+34 91 555 0200",
        "", "ES", "Madrid", "Madrid", "Calle de la Luna 42", "28004",
        "2027-01-31", "SG-IT-Admins;SG-VPN-Access", "SPE_E3;POWER_BI_PRO", "true",
        "support@northwind.com", "false", "H", "", "Contractor engagement until Jan 2027",
    ])
    return buf.getvalue()
